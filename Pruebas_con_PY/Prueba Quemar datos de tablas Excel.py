import openpyxl
import pandas as pd

# Cargar el archivo Excel utilizando openpyxl
wb = openpyxl.load_workbook("Archivo/Libro1.xlsx")
sheet = wb["Hoja1"]

# Leer el contenido de las celdas en un diccionario
data = {}
for row in sheet.iter_rows(values_only=True):
    material = row[0]
    texto = row[1]
    values = row[2:]
    data.setdefault("Material", []).append(material)
    data.setdefault("Texto breve de material", []).append(texto)
    for i, value in enumerate(values, start=1):
        column = f"Columna {i}"
        data.setdefault(column, []).append(value)

# Crear el DataFrame a partir del diccionario
df1 = pd.DataFrame(data)

# Realizar las modificaciones en el DataFrame
df1 = pd.melt(df1, ["Material", "Texto breve de material"])
df1 = df1.dropna(axis=0)

# Guardar el DataFrame en un nuevo archivo Excel utilizando openpyxl
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = "Hoja1"

# Escribir los encabezados de columna en la primera fila
column_headers = list(df1.columns)
for col_num, header in enumerate(column_headers, start=1):
    cell = new_sheet.cell(row=1, column=col_num)
    cell.value = header

# Escribir los datos en las celdas
for row_num, row_values in enumerate(df1.values, start=2):
    for col_num, value in enumerate(row_values, start=1):
        cell = new_sheet.cell(row=row_num, column=col_num)
        cell.value = value

# Guardar el archivo Excel resultante
new_wb.save("Archivo/plan_fab2023.xlsx")
